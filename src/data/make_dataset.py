import os
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from .filtering_function import convert_object_columns_to_integers

def importation_excel(excel_file_path, sheet_name):
    """
    Imports data from a specified Excel file.

    Args:
        excel_file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet in the Excel file.

    Returns:
        pd.DataFrame: Pandas DataFrame containing the data from the specified sheet.
    
    Raises:
        AssertionError: If the Excel file or the specified sheet does not exist.
        AssertionError: If column names are empty.

    Example:
        >>> import_excel_data('my_file.xlsx', 'Sheet1')
    """

    # Check if the Excel file exists
    assert os.path.exists(excel_file_path), f"The Excel file '{excel_file_path}' does not exist."

    # Load the Excel workbook in read-only mode
    workbook = load_workbook(excel_file_path, read_only=True, data_only=True)

    # Check if the specified sheet exists in the Excel file
    assert sheet_name in workbook.sheetnames, f"The sheet '{sheet_name}' does not exist in the Excel file."

    # Select the specified sheet
    sheet = workbook[sheet_name]

    # Get column names from the first row of the sheet
    column_names = [cell.value for cell in sheet[1]]
    # Check if column names are not empty
    assert all(name is not None for name in column_names), "Column names cannot be empty."

    # Get data from the second row of the sheet
    data_row = [cell.value for cell in sheet[2]]

    # Create a Pandas DataFrame with the data and column names
    df = pd.DataFrame([data_row], columns=column_names)

    df = convert_object_columns_to_integers(df)
    # Return the created DataFrame
    return df

def export_excel(proba,excel_file_path,sheets):
    try:

        classeur = openpyxl.load_workbook(excel_file_path,read_only=False,keep_vba=True)

        # Sélectionner la feuille
        feuille_source = classeur[sheets]  # Vous pouvez également spécifier le nom de la feuille, par exemple : classeur['NomDeLaFeuille']

        # Commencer à ajouter les probabilités à la cellule A5
        for index, donnee in enumerate(proba):
            classe = donnee['classe']
            feuille_source.cell(row=5, column=index + 1, value=classe)

        # Commencer à ajouter les probabilités à la cellule A6
        for index, donnee in enumerate(proba):
            probabilite = donnee['probabilite']
            feuille_source.cell(row=6, column=index + 1, value=probabilite)

        # Enregistrer le classeur Excel
        classeur.save(excel_file_path)
        print("Workbook saved successfully.")

    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        # Fermer le classeur Excel
        try:
            classeur.close()
        except:
            pass
